import os
import openpyxl
import pandas as pd
from openpyxl.styles import NamedStyle, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

from src.constantes import FONT_PATH


class Exportador:
    @staticmethod
    def exportar_pdf(lista_jogos, caminho_arquivo):
        if not lista_jogos:
            return False

        doc = SimpleDocTemplate(caminho_arquivo, pagesize=landscape(letter))
        elements = []
        width, height = landscape(letter)
        margin = 0.5 * inch
        effective_width = width - 2 * margin

        data = [
            [
                "Título",
                "Gênero",
                "Plataforma",
                "Data",
                "Estado",
                "Descrição",
                "Tempo",
                "Nota",
            ]
        ]

        for jogo in lista_jogos:
            nota = (
                ""
                if jogo["Forma de Zeramento"] in ["Planejo Jogar", "Desistência"]
                else str(jogo["Nota"])
            )
            data.append(
                [
                    str(jogo["Título"]),
                    str(jogo["Gênero"]),
                    str(jogo["Plataforma"]),
                    str(jogo["Data de Zeramento"]),
                    str(jogo["Forma de Zeramento"]),
                    str(jogo["Descrição de Zeramento"]),
                    str(jogo["Tempo Jogado"]),
                    nota,
                ]
            )

        if os.path.exists(FONT_PATH):
            pdfmetrics.registerFont(TTFont("Mplus1p", FONT_PATH))
            font_name = "Mplus1p"
        else:
            font_name = "Helvetica"

        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            name="TableCell", alignment=1, leading=12, fontName=font_name
        )

        table_data = []
        for row in data:
            table_row = [Paragraph(cell, cell_style) for cell in row]
            table_data.append(table_row)

        col_widths = [effective_width / 8] * 8
        table = Table(table_data, colWidths=col_widths)

        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )

        for i in range(1, len(data)):
            estado = data[i][4]
            if estado == "Planejo Jogar":
                style.add("BACKGROUND", (0, i), (-1, i), colors.beige)
            elif estado == "Desistência":
                style.add("BACKGROUND", (0, i), (-1, i), colors.lightcoral)

        table.setStyle(style)
        elements.append(table)

        try:
            doc.build(elements)
            return True
        except Exception as e:
            print(f"Erro PDF: {e}")
            return False

    @staticmethod
    def exportar_excel(lista_jogos, caminho_arquivo):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Jogos"

        headers = [
            "Título",
            "Gênero",
            "Plataforma",
            "Data de Zeramento",
            "Forma de Zeramento",
            "Descrição",
            "Tempo",
            "Nota",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="333333", end_color="333333", fill_type="solid"
        )

        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        fill_planejo = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        fill_desistencia = PatternFill(
            start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
        )

        for row_idx, jogo in enumerate(lista_jogos, 2):
            valores = [
                jogo["Título"],
                jogo["Gênero"],
                jogo["Plataforma"],
                jogo["Data de Zeramento"],
                jogo["Forma de Zeramento"],
                jogo["Descrição de Zeramento"],
                jogo.get("Tempo Jogado", ""),
                jogo["Nota"],
            ]

            if jogo["Forma de Zeramento"] in ["Planejo Jogar", "Desistência"]:
                valores[7] = ""

            for col_idx, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = Alignment(horizontal="center")

                if jogo["Forma de Zeramento"] == "Planejo Jogar":
                    cell.fill = fill_planejo
                elif jogo["Forma de Zeramento"] == "Desistência":
                    cell.fill = fill_desistencia

        for col in ws.columns:
            column = get_column_letter(col[0].column)
            ws.column_dimensions[column].width = 20

        try:
            wb.save(caminho_arquivo)
            return True
        except Exception as e:
            print(f"Erro Excel: {e}")
            return False

    @staticmethod
    def importar_excel(caminho_arquivo):
        try:
            df = pd.read_excel(caminho_arquivo)
            df = df.fillna("")
            return df.to_dict(orient="records")
        except Exception as e:
            print(f"Erro Importação: {e}")
            return []
