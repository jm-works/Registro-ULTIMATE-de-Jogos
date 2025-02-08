# -------------------------------
# IMPORTS
# -------------------------------

# Módulos do Tkinter
import tkinter as tk
from tkinter import messagebox, filedialog, Menu, END, ttk, font as tkFont, Frame
from tkinter.simpledialog import askinteger

# Manipulação de Dados
import pandas as pd
import numpy as np
import json
import datetime
from datetime import datetime
import re

# Manipulação de Arquivos Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Geração de PDFs
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak, Paragraph
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Criação de Gráficos
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Automação e Sistema
import os
import pyautogui
from screeninfo import get_monitors
import webbrowser
import urllib.parse
import pyperclip

# Manipulação de Imagens
from PIL import Image, ImageTk


# -------------------------------
# FUNÇÕES DE MANIPULAÇÃO DE DADOS
# -------------------------------

def carregar_lista():

    caminho_arquivo = os.path.join("saves", "jogos.json")
    try:
        with open(caminho_arquivo, "r") as arquivo:
            lista_jogos = json.load(arquivo)

            # Ordenar apenas jogos com "Data de Zeramento" válida
            jogos_com_data = [
                jogo for jogo in lista_jogos
                if jogo.get("Data de Zeramento") and re.match(r"^\d{2}/\d{2}/\d{4}$", jogo["Data de Zeramento"])
            ]
            jogos_sem_data = [
                jogo for jogo in lista_jogos
                if not jogo.get("Data de Zeramento") or not re.match(r"^\d{2}/\d{2}/\d{4}$", jogo["Data de Zeramento"])
            ]

            jogos_com_data = sorted(
                jogos_com_data,
                key=lambda jogo: datetime.strptime(
                    jogo["Data de Zeramento"], "%d/%m/%Y")
            )

            # Reunir jogos com e sem data, mantendo os sem data no final
            return jogos_com_data + jogos_sem_data

    except FileNotFoundError:
        return []
    except json.decoder.JSONDecodeError:
        return []


def validar_campos(titulo, genero, plataforma, data_zeramento, tempo_jogado, nota, estado):
    # Validação do campo 'Título'
    if not titulo.strip():
        return "O campo 'Título' é obrigatório! Não deixe seu jogo sem nome."

    # Validação do campo 'Gênero'
    if not genero.strip():
        return "O campo 'Gênero' é obrigatório! Escolha o tipo do seu jogo."

    # Validação do campo 'Plataforma'
    if not plataforma.strip():
        return "O campo 'Plataforma' é obrigatório! Onde você jogou?"

    # Validação do campo 'Forma de Zeramento'
    if not estado.strip():
        return "O campo 'Forma de Zeramento' é obrigatório! Você zerou? Se sim, como?"

    # Validação adicional para estados específicos
    if estado not in ["Planejo Jogar", "Desistência"]:
        # Validação do campo 'Data de Zeramento'
        if not re.match(r'^\d{2}/\d{2}/\d{4}$', data_zeramento):
            return "A data de zeramento deve estar no formato DIA/MÊS/ANO."
        try:
            datetime.strptime(data_zeramento, "%d/%m/%Y")
        except ValueError:
            return "A data de zeramento não é válida!"

        # Validação do campo 'Tempo Jogado'
        if not re.match(r'^\d{1,2}:\d{2}$', tempo_jogado):
            return "O tempo jogado deve estar no formato HORAS:MINUTOS!"
        try:
            horas, minutos = map(int, tempo_jogado.split(":"))
            if horas < 0 or minutos < 0 or minutos >= 60:
                return "O tempo jogado é inválido! Certifique-se de que está no formato correto."
        except ValueError:
            return "O tempo jogado é inválido! Tente novamente."

    # Validação do campo 'Nota'
    if not (1 <= nota <= 10):
        return "A nota deve estar entre 1 e 10!"

    # Tudo está válido
    return None


def organizar_lista(metodo):
    global lista_jogos

    # Separar jogos zerados e não zerados
    jogos_zerados = [jogo for jogo in lista_jogos if jogo.get("Data de Zeramento")]
    jogos_nao_zerados = [jogo for jogo in lista_jogos if not jogo.get("Data de Zeramento")]

    if metodo == "Data de Zeramento":
        jogos_zerados.sort(key=lambda jogo: datetime.strptime(jogo["Data de Zeramento"], "%d/%m/%Y"))
    elif metodo == "Data de Lançamento (Antigo -> Novo)":
        lista_jogos.sort(key=lambda jogo: datetime.strptime(jogo.get("Data de Lançamento", "01/01/3000"), "%d/%m/%Y"))
    elif metodo == "Data de Lançamento (Novo -> Antigo)":
        lista_jogos.sort(key=lambda jogo: datetime.strptime(jogo.get("Data de Lançamento", "01/01/1900"), "%d/%m/%Y"), reverse=True)
    elif metodo == "Ordem Alfabética":
        jogos_zerados.sort(key=lambda jogo: jogo["Título"].lower())
        jogos_nao_zerados.sort(key=lambda jogo: jogo["Título"].lower())
    elif metodo == "Tempo Jogado":
        jogos_zerados.sort(
            key=lambda jogo: int(jogo.get("Tempo Jogado", "0:00").split(":")[0]) * 60 + int(jogo.get("Tempo Jogado", "0:00").split(":")[1]),
            reverse=True
        )
    elif metodo == "Nota":
        jogos_zerados.sort(key=lambda jogo: float(jogo.get("Nota", 0)), reverse=True)
    elif metodo == "Plataforma":
        jogos_zerados.sort(key=lambda jogo: (jogo["Plataforma"].lower(), jogo["Título"].lower()))
        jogos_nao_zerados.sort(key=lambda jogo: (jogo["Plataforma"].lower(), jogo["Título"].lower()))

    # Unindo novamente os jogos zerados e não zerados, mantendo os não zerados no final
    lista_jogos = jogos_zerados + jogos_nao_zerados

    atualizar_lista(lista_jogos)

def abrir_menu_organizacao():
    janela_organizacao = tk.Toplevel(root)
    janela_organizacao.title("Organizar Jogos")
    janela_organizacao.resizable(False, False)
    janela_organizacao.geometry("300x270")
    centralizar_janela(janela_organizacao, 300, 270)
    
    tk.Label(janela_organizacao, text="Escolha o método de organização:", font=("Arial", 10, "bold")).pack(pady=10)
    
    metodos = ["Data de Zeramento", "Ordem Alfabética", "Tempo Jogado", "Nota", "Plataforma"]
    for metodo in metodos:
        tk.Button(janela_organizacao, text=metodo, command=lambda m=metodo: [organizar_lista(m), janela_organizacao.destroy()]).pack(pady=5)


def atualizar_entry(event):
    genero_entry.delete(0, tk.END)
    texto_digitado = genero_combobox.get()
    genero_entry.insert(0, texto_digitado)


def limpar_filtros():
    global jogos_filtrados
    jogos_filtrados = lista_jogos.copy()
    atualizar_lista(jogos_filtrados)
    if filtro_window:
        filtro_window.destroy()


def limpar_campos():
    global nota_entry, titulo_entry, genero_entry, plataforma_entry, data_zeramento_entry, descricao_zeramento_entry, tempo_jogado_entry

    if nota_entry:
        nota_entry.delete(0, tk.END)
    if titulo_entry:
        titulo_entry.delete(0, tk.END)
    if genero_entry:
        genero_entry.delete(0, tk.END)
    if plataforma_entry:
        plataforma_entry.delete(0, tk.END)
    if data_zeramento_entry:
        data_zeramento_entry.delete(0, tk.END)
    if descricao_zeramento_entry:
        descricao_zeramento_entry.delete(0, tk.END)
    if tempo_jogado_entry:
        tempo_jogado_entry.delete(0, tk.END)

    # Limpar as combobox
    genero_combobox.set('')
    plataforma_combobox.set('')
    forma_zeramento_combobox.set('')


def formatar_tempo_jogado(event):
    texto = tempo_jogado_entry.get()
    # Remove caracteres não numéricos
    texto = ''.join(filter(str.isdigit, texto))

    # Formatação automática
    if len(texto) > 2:
        texto = texto[:2] + ":" + texto[2:]

    # Limitar a dois dígitos para horas e minutos
    if len(texto) > 5:
        texto = texto[:5]

    tempo_jogado_entry.delete(0, tk.END)
    tempo_jogado_entry.insert(0, texto)


def formatar_data(entry):
    input_text = entry.get()

    input_text = ''.join(filter(str.isdigit, input_text))

    if len(input_text) >= 2:
        formatted_text = input_text[:2]
        if len(input_text) >= 4:
            formatted_text += '/' + input_text[2:4]
            if len(input_text) >= 8:
                formatted_text += '/' + input_text[4:8]
        entry.delete(0, tk.END)
        entry.insert(0, formatted_text)


def centralizar_janela(janela, largura, altura):
    janela.update_idletasks()
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()
    x = (largura_tela - largura) // 2
    y = (altura_tela - altura) // 2
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


def formatar_data_zeramento(event):
    input_text = data_zeramento_entry.get()

    # Remover todos os caracteres que não sejam dígitos ou parênteses
    formatted_text = re.sub(r'[^0-9()]', '', input_text)

    # Limitar o ano a 4 dígitos
    if '/' in formatted_text:
        parts = formatted_text.split('/')
        if len(parts) > 2:
            parts[2] = parts[2][:4]
            formatted_text = '/'.join(parts)

    # Recriar a entrada formatada, adicionando barras automaticamente
    if len(formatted_text) >= 2 and formatted_text[1] != '/':
        formatted_text = formatted_text[:2] + '/' + formatted_text[2:]
    if len(formatted_text) >= 5 and formatted_text[4] != '/':
        formatted_text = formatted_text[:5] + '/' + formatted_text[5:]

    # Limitar o ano a 4 dígitos
    if '/' in formatted_text:
        parts = formatted_text.split('/')
        if len(parts) > 2:
            parts[2] = parts[2][:4]
            formatted_text = '/'.join(parts)

    data_zeramento_entry.delete(0, tk.END)
    data_zeramento_entry.insert(0, formatted_text)


def is_valid_nota(nota):
    try:
        nota = float(nota)
        return 1 <= nota <= 10
    except ValueError:
        return False


def atualizar_data_zeramento(event):
    data_atual = datetime.now().strftime("%d/%m/%Y")
    data_zeramento_entry.delete(0, tk.END)
    data_zeramento_entry.insert(0, data_atual)
    formatar_data_zeramento(data_atual)


def excluir_jogo():
    selecionado = lista_jogos_listbox.curselection()
    if not selecionado:
        messagebox.showerror("Erro", "Selecione um jogo para excluir.")
        return

    indice = int(selecionado[0])

    if 0 <= indice < len(lista_jogos):
        # Confirmar exclusão
        jogo_a_excluir = lista_jogos[indice]
        confirmar = messagebox.askyesno(
            "Confirmação", f"Tem certeza que deseja excluir o jogo '{jogo_a_excluir['Título']}'?"
        )

        if confirmar:
            jogo_excluido = lista_jogos.pop(indice)
            atualizar_lista()
            messagebox.showinfo(
                "Sucesso", f"Jogo '{jogo_excluido['Título']}' excluído com sucesso!")
            if janela_edicao:
                janela_edicao.destroy()


def adicionar_jogo():
    # Capturar os valores dos campos
    titulo = titulo_entry.get()
    genero = genero_entry.get()
    plataforma = plataforma_entry.get()
    data_zeramento = data_zeramento_entry.get()
    tempo_jogado = tempo_jogado_entry.get()
    nota = nota_slider.get()
    estado = forma_zeramento_combobox.get()

    # Validar os campos
    mensagem_erro = validar_campos(
        titulo, genero, plataforma, data_zeramento, tempo_jogado, nota, estado)
    if mensagem_erro:
        show_error_message(mensagem_erro)
        return

    # Criar o novo jogo
    novo_jogo = {
        "Título": titulo,
        "Gênero": genero,
        "Plataforma": plataforma,
        "Data de Zeramento": data_zeramento if estado not in ["Planejo Jogar", "Desistência"] else "",
        "Forma de Zeramento": estado,
        "Descrição de Zeramento": descricao_zeramento_entry.get(),
        "Tempo Jogado": tempo_jogado if estado not in ["Planejo Jogar", "Desistência"] else "",
        "Nota": nota,
    }

    # Adicionar o jogo à lista
    lista_jogos.append(novo_jogo)
    atualizar_lista()
    limpar_filtros()
    limpar_campos()

    # Mensagem de sucesso
    messagebox.showinfo(
        "Sucesso!", f"O jogo '{titulo}' foi adicionado com sucesso!")


def atualizar_campos(event):
    estado = forma_zeramento_combobox.get()
    if estado in ["Planejo Jogar", "Desistência"]:
        # Desativar campos
        tempo_jogado_entry_var.set("")  # Use set() para limpar o campo
        tempo_jogado_entry.config(state="disabled")
        nota_slider.set(1)
        nota_slider.config(state="disabled")
        data_zeramento_entry.delete(0, tk.END)
        data_zeramento_entry.config(state="disabled")
    else:
        # Reativar campos
        tempo_jogado_entry.config(state="normal")
        nota_slider.config(state="normal")
        data_zeramento_entry.config(state="normal")


def atualizar_genero(event):
    genero_entry.delete(0, tk.END)
    genero_selecionado = genero_combobox.get()
    genero_entry.insert(0, genero_selecionado)


def atualizar_plataforma(event):
    plataforma_entry.delete(0, tk.END)
    plataforma_entry.insert(0, plataforma_var.get())


def validar_numero(P):
    # Esta função permite entrada numérica e espaços
    if P == "" or re.match("^[0-9\s]*$", P):
        return True
    else:
        return False


def atualizar_lista(jogos=None):
    if jogos is None:
        jogos = lista_jogos
    lista_jogos_listbox.delete(0, tk.END)
    for idx, jogo in enumerate(jogos, start=1):
        estado = jogo["Forma de Zeramento"]
        if estado == "História":
            icone = "✅"
        elif estado == "Planejo Jogar":
            icone = "📅"
        elif estado == "Desistência":
            icone = "❌"
        else:
            icone = "✅"
        lista_jogos_listbox.insert(tk.END, f"{idx}. {icone} {jogo['Título']}")


# -------------------------------
# FUNÇÕES DE IMPORTAÇÃO & EXPORTAÇÃO
# -------------------------------

def exportar_para_pdf():
    if not lista_jogos:
        messagebox.showerror(
            "Erro", "A lista de jogos está vazia. Não há nada para exportar.")
        return

    nome_arquivo = filedialog.asksaveasfilename(
        defaultextension=".pdf", filetypes=[("Arquivos PDF", "*.pdf")])
    if nome_arquivo:
        # Criar um documento PDF
        doc = SimpleDocTemplate(nome_arquivo, pagesize=landscape(letter))
        elements = []

        # Calcular a largura da página e margens
        width, height = landscape(letter)
        margin = 0.5 * inch  # Margem de 0,5 polegadas nas laterais

        # Calcular a largura efetiva da tabela
        effective_width = width - 2 * margin

        # Criar uma tabela para os dados da lista de jogos
        data = [["Título", "Gênero", "Plataforma", "Data de Zeramento",
                 "Forma de Zeramento", "Descrição de Zeramento", "Tempo Jogado", "Nota"]]
        for jogo in lista_jogos:
            # Remover a nota para "Planejo Jogar" e "Desistência"
            nota = "" if jogo["Forma de Zeramento"] in [
                "Planejo Jogar", "Desistência"] else str(jogo["Nota"])
            data.append([str(jogo["Título"]), str(jogo["Gênero"]), str(jogo["Plataforma"]), str(jogo["Data de Zeramento"]),
                         str(jogo["Forma de Zeramento"]), str(jogo["Descrição de Zeramento"]), str(jogo["Tempo Jogado"]), nota])

        # Calcular a largura das colunas com base no conteúdo, limitando a um valor máximo
        max_col_width = effective_width / len(data[0])
        colWidths = [min(max(len(data[i][j]) * 7, max_col_width)
                         for i in range(len(data))) for j in range(len(data[0]))]

        # Definir um estilo de parágrafo para o conteúdo da célula
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(name='TableCell')
        cell_style.alignment = 1  # Centralizar o texto na célula
        cell_style.leading = 12  # Espaçamento entre linhas

        # Adicionar a tabela de dados com cores por estado
        table_data = []
        for i, row in enumerate(data):
            table_row = []
            for cell_content in row:
                # Criar um parágrafo com o estilo da célula
                cell_paragraph = Paragraph(cell_content, cell_style)
                table_row.append(cell_paragraph)
            table_data.append(table_row)

        # Criar a tabela e definir o estilo
        table = Table(table_data, colWidths=colWidths)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centralizar verticalmente
        ])

        # Adicionar cores específicas por estado
        for i in range(1, len(data)):
            estado = data[i][4]  # Índice da coluna "Forma de Zeramento"
            if estado == "Planejo Jogar":
                style.add('BACKGROUND', (0, i), (-1, i),
                          colors.beige)  # Cor bege
            elif estado == "Desistência":
                style.add('BACKGROUND', (0, i), (-1, i),
                          colors.lightcoral)  # Cor vermelho claro
            else:
                style.add('BACKGROUND', (0, i), (-1, i),
                          colors.white)  # Cor branca padrão

        # Aplicar estilo à tabela
        table.setStyle(style)

        # Adicionar a tabela ao elemento PDF
        elements.append(Spacer(1, 0.25 * inch))  # Espaço antes da tabela
        elements.append(table)
        elements.append(Spacer(1, 0.25 * inch))  # Espaço após a tabela

        # Construir o PDF
        doc.build(elements)

        messagebox.showinfo(
            "Sucesso", f"Lista de jogos exportada para {nome_arquivo} (PDF) com sucesso!")


def exportar_para_excel():
    if not lista_jogos:
        messagebox.showerror(
            "Erro", "A lista de jogos está vazia. Não há nada para exportar.")
        return

    # Criar um novo arquivo Excel e uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Jogos"

    # Definir estilos para o cabeçalho
    header_style = NamedStyle(name="header_style")
    header_style.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_style.fill = openpyxl.styles.PatternFill(
        start_color="333333", end_color="333333", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center")  # Centralizar
    for col_num, column_title in enumerate(["Título", "Gênero", "Plataforma", "Data de Zeramento",
                                            "Forma de Zeramento", "Descrição de Zeramento", "Tempo Jogado", "Nota"], 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        # Atribuir diretamente o estilo ao cabeçalho
        cell.style = header_style

    # Definir estilos de cores para "Planejo Jogar" e "Desistência"
    estilo_planejo = openpyxl.styles.PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Bege
    estilo_desistencia = openpyxl.styles.PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Vermelho claro

    # Preencher os dados
    for row_num, jogo in enumerate(lista_jogos, start=2):
        for col_num, (chave, valor) in enumerate(jogo.items(), start=1):
            # Ajustar a nota para "Planejo Jogar" e "Desistência"
            if chave == "Nota" and jogo["Forma de Zeramento"] in ["Planejo Jogar", "Desistência"]:
                valor = ""
            cell = ws.cell(row=row_num, column=col_num, value=valor)

            # Aplicar cores conforme o estado do jogo
            if jogo["Forma de Zeramento"] == "Planejo Jogar":
                cell.fill = estilo_planejo
            elif jogo["Forma de Zeramento"] == "Desistência":
                cell.fill = estilo_desistencia

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(horizontal="center")

    # Salvar o arquivo Excel
    nome_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
    if nome_arquivo:
        wb.save(nome_arquivo)
        messagebox.showinfo(
            "Sucesso", f"Lista de jogos exportada para {nome_arquivo} com sucesso!")


def importar_de_excel():
    nome_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx")])
    if not nome_arquivo:
        return  # Cancelamento do diálogo

    try:
        # Ler o arquivo Excel em um DataFrame
        df = pd.read_excel(nome_arquivo)

        # Verificar se as colunas necessárias estão presentes
        colunas_esperadas = ["Título", "Gênero", "Plataforma", "Data de Zeramento",
                             "Forma de Zeramento", "Descrição de Zeramento", "Tempo Jogado", "Nota"]
        for coluna in colunas_esperadas:
            if coluna not in df.columns:
                messagebox.showerror(
                    "Erro", f"A coluna '{coluna}' está ausente no arquivo Excel.")
                return

        # Substituir valores NaN por strings vazias
        df = df.fillna("")

        # Converter o DataFrame em uma lista de dicionários
        lista_importada = df.to_dict(orient='records')

        # Criar um conjunto de identificadores únicos (título + plataforma + estado)
        jogos_existentes = {
            (jogo["Título"].strip().lower(), jogo["Plataforma"].strip(
            ).lower(), jogo["Forma de Zeramento"].strip().lower())
            for jogo in lista_jogos
        }

        # Adicionar todos os jogos (incluindo validação de campos obrigatórios)
        jogos_adicionados = 0
        for jogo in lista_importada:
            # Criar um identificador único para o jogo
            identificador = (
                jogo["Título"].strip().lower(),
                jogo["Plataforma"].strip().lower(),
                jogo["Forma de Zeramento"].strip().lower()
            )

            # Verificar título vazio ou inválido
            if not jogo["Título"].strip():
                print(f"Jogo ignorado por título vazio: {jogo}")
                continue  # Ignorar jogos sem título

            # Verificar se o jogo já existe na lista
            if identificador in jogos_existentes:
                print(f"Jogo ignorado por ser duplicado: {jogo}")
                continue

            # Validar e formatar a nota
            try:
                nota = float(jogo["Nota"])
                if 1 <= nota <= 10:
                    jogo["Nota"] = nota
                else:
                    jogo["Nota"] = ""  # Nota inválida
            except (ValueError, TypeError):
                jogo["Nota"] = ""  # Nota ausente ou inválida

            # Adicionar o jogo à lista
            lista_jogos.append(jogo)
            # Atualizar o conjunto de identificadores
            jogos_existentes.add(identificador)
            jogos_adicionados += 1

        # Atualizar a lista exibida
        atualizar_lista()
        limpar_filtros()

        # Mensagem de sucesso
        if jogos_adicionados > 0:
            messagebox.showinfo(
                "Sucesso", f"{jogos_adicionados} jogos foram importados de {nome_arquivo} com sucesso!"
            )
        else:
            messagebox.showinfo(
                "Informação", "Nenhum novo jogo foi importado. Todos os jogos já existem na lista."
            )

    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo não encontrado.")
    except ValueError as e:
        messagebox.showerror("Erro", f"Erro ao processar o arquivo Excel: {e}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro inesperado: {e}")


def salvar_lista():
    os.makedirs("saves", exist_ok=True)
    caminho_arquivo = os.path.join("saves", "jogos.json")
    with open(caminho_arquivo, "w") as arquivo:
        json.dump(lista_jogos, arquivo)
    messagebox.showinfo("Salvar", "Lista de jogos salva com sucesso!")

# -------------------------------
# FUNÇÕES DE GERAÇÃO DE ESTATISTICAS
# -------------------------------


def criar_distribuicao_plataformas():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    # Filtrar jogos válidos (excluir "Planejo Jogar" e "Desistência")
    jogos_validos = [
        jogo for jogo in lista_jogos
        if jogo.get("Forma de Zeramento") not in ["Planejo Jogar", "Desistência"]
    ]

    # Contagem de jogos por plataforma
    plataforma_contagem = {}
    for jogo in jogos_validos:
        plataforma = jogo["Plataforma"]
        plataforma_contagem[plataforma] = plataforma_contagem.get(
            plataforma, 0) + 1

    if not plataforma_contagem:
        messagebox.showinfo(
            "Informação", "Não há jogos válidos para exibir no gráfico.")
        return

    plataformas = list(plataforma_contagem.keys())
    contagens = list(plataforma_contagem.values())

    # Adicionar os números nos rótulos das plataformas
    labels = [f"{plataforma} ({quantidade})" for plataforma,
              quantidade in zip(plataformas, contagens)]

    # Criar o gráfico de pizza
    plt.figure(figsize=(8, 6))
    plt.pie(contagens, labels=labels, autopct='%1.1f%%',
            startangle=140, colors=plt.cm.Paired.colors)

    # Configurações do gráfico
    plt.title("Jogos Zerados por Plataforma",
              fontsize=14, fontweight='bold')
    plt.axis('equal')  # Garantir que o gráfico seja um círculo
    plt.tight_layout()

    # Exibir o gráfico
    plt.show()


def criar_media_notas_plataformas():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    # Filtrar jogos válidos (excluir "Planejo Jogar" e "Desistência")
    jogos_validos = [
        jogo for jogo in lista_jogos
        if jogo.get("Forma de Zeramento") not in ["Planejo Jogar", "Desistência"]
    ]

    if not jogos_validos:
        messagebox.showinfo(
            "Informação", "Não há jogos válidos para calcular a média de notas.")
        return

    # Calcular médias por plataforma
    notas_por_plataforma = {}
    for jogo in jogos_validos:
        plataforma = jogo["Plataforma"]
        nota = float(jogo["Nota"])
        if plataforma not in notas_por_plataforma:
            notas_por_plataforma[plataforma] = []
        notas_por_plataforma[plataforma].append(nota)

    # Calcular a média e preparar os rótulos
    plataformas = list(notas_por_plataforma.keys())
    medias = [sum(notas) / len(notas)
              for notas in notas_por_plataforma.values()]
    labels = [
        f"{plataforma} ({len(notas_por_plataforma[plataforma])})" for plataforma in plataformas]

    # Gráfico de barras
    plt.figure(figsize=(10, 6))
    plt.bar(plataformas, medias, color='skyblue', edgecolor='black')

    # Adicionar valores acima das barras
    for i, media in enumerate(medias):
        plt.text(i, media + 0.1, f"{media:.2f}", ha='center', fontsize=10)

    # Configurações do gráfico
    plt.title("Média de Notas por Plataforma",
              fontsize=14, fontweight='bold')
    plt.xlabel("Plataformas", fontsize=12)
    plt.ylabel("Média das Notas", fontsize=12)
    plt.xticks(range(len(plataformas)), labels, rotation=45, fontsize=10)
    plt.tight_layout()

    # Exibir o gráfico
    plt.show()


def criar_tempo_total_plataformas():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    # Filtrar jogos válidos (excluir "Planejo Jogar" e "Desistência")
    jogos_validos = [
        jogo for jogo in lista_jogos
        if jogo.get("Forma de Zeramento") not in ["Planejo Jogar", "Desistência"]
    ]

    if not jogos_validos:
        messagebox.showinfo(
            "Informação", "Não há jogos válidos para calcular o tempo total.")
        return

    # Calcular tempo total por plataforma
    tempo_por_plataforma = {}
    for jogo in jogos_validos:
        plataforma = jogo["Plataforma"]
        tempo_jogado = jogo.get("Tempo Jogado", "").strip()

        # Verificar se o tempo jogado está no formato "HH:MM"
        if re.match(r"^\d{1,2}:\d{2}$", tempo_jogado):
            horas, minutos = map(int, tempo_jogado.split(":"))
            tempo_total = horas * 60 + minutos
            tempo_por_plataforma[plataforma] = tempo_por_plataforma.get(
                plataforma, 0) + tempo_total
        else:
            # Ignorar jogos com tempo inválido
            continue

    if not tempo_por_plataforma:
        messagebox.showinfo(
            "Informação", "Não há tempos válidos para exibir no gráfico.")
        return

    plataformas = list(tempo_por_plataforma.keys())
    # Converter minutos para horas
    tempos_horas = [tempo // 60 for tempo in tempo_por_plataforma.values()]
    labels = [f"{plataforma} ({tempos_horas[i]}h)" for i,
              plataforma in enumerate(plataformas)]

    # Gráfico de barras
    plt.figure(figsize=(10, 6))
    plt.bar(plataformas, tempos_horas, color='lightgreen', edgecolor='black')

    # Adicionar valores acima das barras
    for i, tempo in enumerate(tempos_horas):
        plt.text(i, tempo + 0.1, f"{tempo}h", ha='center', fontsize=10)

    # Configurações do gráfico
    plt.title("Tempo Total de zeramento por Plataforma",
              fontsize=14, fontweight='bold')
    plt.xlabel("Plataformas", fontsize=12)
    plt.ylabel("Tempo Total (Horas)", fontsize=12)
    plt.xticks(range(len(plataformas)), labels, rotation=45, fontsize=10)
    plt.tight_layout()

    # Exibir o gráfico
    plt.show()


def criar_grafico_jogos_por_ano():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    jogos_por_ano = {}

    for jogo in lista_jogos:
        try:
            ano = datetime.strptime(jogo["Data de Zeramento"], "%d/%m/%Y").year
            jogos_por_ano[ano] = jogos_por_ano.get(ano, 0) + 1
        except ValueError:
            continue

    anos = sorted(jogos_por_ano.keys())
    contagem = [jogos_por_ano[ano] for ano in anos]

    plt.figure(figsize=(8, 5))
    plt.bar(anos, contagem, color='skyblue')
    plt.xlabel("Ano")
    plt.ylabel("Quantidade de Jogos Zerados")
    plt.title("Jogos Zerados por Ano")
    plt.xticks(anos, rotation=45)
    plt.tight_layout()
    plt.show()


def criar_grafico_comparativo_generos():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    # Filtrar apenas os jogos zerados (com Data de Zeramento válida)
    jogos_zerados = [
        jogo for jogo in lista_jogos if jogo.get("Data de Zeramento")]

    # Dicionário para armazenar os dados por ano e gênero
    dados_generos = {}

    for jogo in jogos_zerados:
        try:
            ano = datetime.strptime(jogo["Data de Zeramento"], "%d/%m/%Y").year
            genero = jogo["Gênero"]
            if ano not in dados_generos:
                dados_generos[ano] = {}
            dados_generos[ano][genero] = dados_generos[ano].get(genero, 0) + 1
        except ValueError:
            continue

    # Preparar os dados para o gráfico
    anos = sorted(dados_generos.keys())
    generos = sorted(set(g for dados in dados_generos.values() for g in dados))

    valores = {genero: [dados_generos.get(ano, {}).get(
        genero, 0) for ano in anos] for genero in generos}

    # Plotar o gráfico
    plt.figure(figsize=(10, 6))
    for genero, valores_genero in valores.items():
        plt.plot(anos, valores_genero, marker='o', label=genero)

    # Configurações do gráfico
    plt.xlabel("Ano", fontsize=12)
    plt.ylabel("Quantidade de Jogos Zerados", fontsize=12)
    plt.title("Comparação de Gêneros Zerados por Ano",
              fontsize=14, fontweight="bold")
    plt.xticks(anos, rotation=45, fontsize=10)
    plt.yticks(fontsize=10)
    plt.legend(title="Gêneros", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    plt.show()


def criar_analise_de_notas():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    # Filtrar jogos válidos para análise (excluir "Planejo Jogar" e "Desistência")
    jogos_validos = [
        jogo for jogo in lista_jogos
        if jogo.get("Forma de Zeramento") not in ["Planejo Jogar", "Desistência"]
    ]

    # Calcular contagem de jogos por nota
    notas = [int(jogo["Nota"]) for jogo in jogos_validos if jogo["Nota"]]
    if not notas:
        messagebox.showinfo(
            "Informação", "Não há jogos válidos para análise de notas.")
        return

    contagem_notas = {i: notas.count(i) for i in range(1, 11)}

    # Calcular a média das notas
    media_notas = np.mean(notas)

    # Gráfico de barras
    plt.figure(figsize=(10, 6))
    plt.bar(contagem_notas.keys(), contagem_notas.values(),
            color='skyblue', edgecolor='black')

    # Destacar a média das notas
    plt.axhline(y=media_notas, color='red', linestyle='--',
                label=f'Média: {media_notas:.2f}')

    # Configurações do gráfico
    plt.title("Análise de Notas",
              fontsize=14, fontweight='bold')
    plt.xlabel("Notas", fontsize=12)
    plt.ylabel("Quantidade de Jogos", fontsize=12)
    plt.xticks(range(1, 11), fontsize=10)
    plt.yticks(fontsize=10)
    plt.legend(loc="upper right", fontsize=10)
    plt.tight_layout()

    # Exibir o gráfico
    plt.show()


def contar_jogos_zerados_por_ano(ano):
    numero_jogos_zerados = 0
    for jogo in lista_jogos:
        data_zeramento_str = jogo["Data de Zeramento"]
        if data_zeramento_str:
            data_zeramento = datetime.strptime(data_zeramento_str, "%d/%m/%Y")
            if data_zeramento.year == ano:
                numero_jogos_zerados += 1
    return numero_jogos_zerados


def calcular_tempo_total_jogado():
    total_minutos = 0

    for jogo in lista_jogos:
        tempo_jogado = jogo.get("Tempo Jogado", "").strip()

        # Validar se o tempo jogado está no formato correto (HH:MM)
        if re.match(r"^\d{1,2}:\d{2}$", tempo_jogado):
            horas, minutos = map(int, tempo_jogado.split(":"))
            total_minutos += horas * 60 + minutos
        else:
            # Ignorar jogos com tempo inválido
            continue

    total_horas = total_minutos // 60
    total_minutos = total_minutos % 60
    total_dias = total_horas // 24
    total_horas = total_horas % 24

    messagebox.showinfo(
        "Tempo Total Jogado", f"Você perdeu {total_dias} dias, {total_horas} horas, {total_minutos} minutos, da sua vida com jogos"
    )


def calcular_total_minutos(tempo_jogado):
    horas, minutos = map(int, tempo_jogado.split(":"))
    return horas * 60 + minutos


def criar_grafico_generos():
    if not lista_jogos:
        messagebox.showinfo("Informação", "A lista de jogos está vazia.")
        return

    genero_contagem = {}

    for jogo in lista_jogos:
        estado = jogo["Forma de Zeramento"]
        # Considerar apenas jogos "completos" ou "desistidos"
        if estado in ["História", "100%", "Platina", "Desistência"]:
            genero = jogo["Gênero"]
            if genero in genero_contagem:
                genero_contagem[genero] += 1
            else:
                genero_contagem[genero] = 1

    if not genero_contagem:
        messagebox.showinfo(
            "Informação", "Não há jogos válidos para exibir no gráfico.")
        return

    # Preparar dados do gráfico
    generos = list(genero_contagem.keys())
    contagem = list(genero_contagem.values())

    # Adicionar números nos rótulos
    labels = [f"{genero} ({quantidade})" for genero,
              quantidade in zip(generos, contagem)]

    # Criar gráfico de pizza
    plt.figure(figsize=(8, 8))
    plt.pie(contagem, labels=labels, autopct='%1.1f%%', startangle=140)
    plt.title('Distribuição de Gêneros (Número de Jogos)')
    plt.axis('equal')  # Garantir que o gráfico seja um círculo
    plt.show()


def exibir_numero_jogos_zerados_por_ano():
    try:
        ano = askinteger("Jogos Zerados (Por ano)", "Digite o ano desejado:")
        if ano is not None:
            numero_jogos_zerados = contar_jogos_zerados_por_ano(ano)
            messagebox.showinfo(
                "Número de Jogos Zerados", f"Foram zerados {numero_jogos_zerados} jogos em {ano}.")
    except ValueError:
        messagebox.showerror("Erro", "Ano inválido. Digite um ano válido.")


def mostrar_jogos_longos_curto():
    # Filtrar apenas os jogos zerados com tempo jogado válido
    jogos_zerados = [
        jogo for jogo in lista_jogos
        if jogo.get("Data de Zeramento") and jogo.get("Tempo Jogado")
    ]

    if not jogos_zerados:
        messagebox.showinfo("Informação", "Não há jogos zerados para exibir.")
        return

    try:
        # Ordenar os jogos por tempo jogado
        jogos_zerados.sort(
            key=lambda jogo: calcular_total_minutos(jogo["Tempo Jogado"]))

        # Identificar o mais curto e o mais longo
        jogo_mais_curto = jogos_zerados[0]
        jogo_mais_longo = jogos_zerados[-1]

        # Exibir os resultados
        mensagem = (
            f"Jogo mais curto:\n"
            f"{jogo_mais_curto['Título']} - {jogo_mais_curto['Tempo Jogado']}\n\n"
            f"Jogo mais longo:\n"
            f"{jogo_mais_longo['Título']} - {jogo_mais_longo['Tempo Jogado']}"
        )
        messagebox.showinfo("Jogos mais e Menos longos", mensagem)
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")


def salvar_em_arquivo():
    nome_arquivo = filedialog.asksaveasfilename(
        defaultextension=".json", filetypes=[("Arquivos JSON", "*.json")])
    if nome_arquivo:
        with open(nome_arquivo, "w") as arquivo:
            json.dump(lista_jogos, arquivo)
        messagebox.showinfo(
            "Salvar em Arquivo", f"Lista de jogos salva em {nome_arquivo} com sucesso!")


def carregar_de_arquivo():
    nome_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos JSON", "*.json")])
    if nome_arquivo:
        try:
            with open(nome_arquivo, "r") as arquivo:
                lista_carregada = json.load(arquivo)
            lista_jogos.extend(lista_carregada)
            atualizar_lista()
            messagebox.showinfo(
                "Carregar de Arquivo", f"Lista de jogos carregada de {nome_arquivo} com sucesso!")
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado.")
        except Exception as e:
            messagebox.showerror(
                "Erro", f"Ocorreu um erro ao carregar o arquivo JSON: {e}")

# -------------------------------
# FUNÇÕES DE INTERFACE GRÁFICA (UI)
# -------------------------------


def estilizar_botao(botao, cor_fundo, cor_texto="white", fonte=("Arial", 10, "bold"),
                    largura=20, altura=2, wrap=150, borda=3, efeito_hover="#333333"):
    botao.config(
        bg=cor_fundo, fg=cor_texto,
        font=fonte,
        width=largura, height=altura,
        wraplength=wrap,
        relief="raised", bd=borda,
        cursor="hand2"
    )

    # Efeito hover personalizável
    botao.bind("<Enter>", lambda e: botao.config(bg=efeito_hover))
    botao.bind("<Leave>", lambda e: botao.config(bg=cor_fundo))


def abrir_menu_contexto(event):
    # Obter o índice do item clicado com o botão direito
    index = lista_jogos_listbox.nearest(event.y)

    # Selecionar o item correspondente
    lista_jogos_listbox.selection_clear(0, END)  # Limpar seleções anteriores
    lista_jogos_listbox.selection_set(index)  # Selecionar o item clicado
    lista_jogos_listbox.activate(index)  # Focar no item clicado

    # Verificar se há uma seleção válida antes de abrir o menu
    if not lista_jogos_listbox.curselection():
        return

    # Criar um menu de contexto
    menu = Menu(lista_jogos_listbox, tearoff=0)
    menu.add_command(label="Pesquisar no Google", command=pesquisar_no_google)
    menu.add_command(label="Copiar nome do Jogo", command=copiar_nome)
    menu.add_command(label="Organizar lista", command=abrir_menu_organizacao)
    menu.add_command(label="Filtrar", command=mostrar_jogos_filtrados)
    menu.add_command(label="Limpar Filtros", command=limpar_filtros)
    menu.add_command(label="Editar jogo", command=editar_jogo)
    menu.add_command(label="Excluir Jogo", command=excluir_jogo)

    # Exibir o menu na posição do cursor
    menu.post(event.x_root, event.y_root)


def pesquisar_no_google():
    # Obtenha o índice do jogo selecionado
    indice_selecionado = lista_jogos_listbox.curselection()[0]
    jogo_selecionado = lista_jogos_listbox.get(indice_selecionado)

    # Use uma expressão regular para encontrar o título do jogo após o número e o ponto
    padrao_titulo = r'\d+\.\s*(.+)'
    correspondencia = re.search(padrao_titulo, jogo_selecionado)
    if correspondencia:
        titulo_do_jogo = correspondencia.group(1)
    else:
        titulo_do_jogo = "Título não encontrado"

    # Remover apenas os emojis ✅, 📅, ❌
    titulo_sem_emojis = re.sub(r'[✅📅❌]', '', titulo_do_jogo)

    # Criar uma URL de pesquisa no Google com o título limpo
    pesquisa_google_url = f"https://www.google.com/search?q={urllib.parse.quote(titulo_sem_emojis)}"

    # Abrir o navegador padrão para realizar a pesquisa
    webbrowser.open(pesquisa_google_url)


def editar_jogo():
    global filtro_window, janela_edicao, nota_entry
    selecionado = lista_jogos_listbox.curselection()
    if not selecionado:
        messagebox.showerror("Erro", "Selecione um jogo para editar.")
        return

    indice = int(selecionado[0])

    if 0 <= indice < len(lista_jogos):
        jogo_selecionado = lista_jogos[indice]

        janela_edicao = tk.Toplevel(root)
        janela_edicao.title("Editar Jogo")
        janela_edicao.geometry("330x370")
        janela_edicao.resizable(False, False)
        centralizar_janela(janela_edicao, 330, 370)

        # Criar rótulos e entradas
        tk.Label(janela_edicao, text="Título:", font=("Arial", 10, "bold")).grid(
            row=0, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Gênero:", font=("Arial", 10, "bold")).grid(
            row=1, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Plataforma:", font=("Arial", 10, "bold")).grid(
            row=2, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Data de Zeramento:", font=(
            "Arial", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Forma de Zeramento:", font=(
            "Arial", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Descrição de Zeramento:", font=(
            "Arial", 10, "bold")).grid(row=5, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Tempo Jogado (HH:MM):", font=(
            "Arial", 10, "bold")).grid(row=6, column=0, padx=5, pady=5, sticky="W")
        tk.Label(janela_edicao, text="Nota:", font=("Arial", 10, "bold")).grid(
            row=7, column=0, padx=5, pady=5, sticky="W")

        # Criar widgets de entrada
        titulo_edit = tk.Entry(janela_edicao)
        titulo_edit.insert(0, jogo_selecionado['Título'])

        genero_edit = ttk.Combobox(
            janela_edicao, values=generos_disponiveis, state="readonly")
        genero_edit.set(jogo_selecionado['Gênero'])

        plataforma_edit = ttk.Combobox(
            janela_edicao, values=plataformas_disponiveis, state="readonly")
        plataforma_edit.set(jogo_selecionado['Plataforma'])

        data_zeramento_edit = tk.Entry(janela_edicao)
        data_zeramento_edit.insert(0, jogo_selecionado['Data de Zeramento'])

        forma_zeramento_edit = ttk.Combobox(janela_edicao, values=[
                                            "História", "100%", "Platina", "Planejo Jogar", "Desistência", "Outro"], state="readonly")
        forma_zeramento_edit.set(jogo_selecionado['Forma de Zeramento'])

        descricao_zeramento_edit = tk.Entry(janela_edicao)
        descricao_zeramento_edit.insert(
            0, jogo_selecionado['Descrição de Zeramento'])

        tempo_jogado_edit_horas = tk.Spinbox(
            janela_edicao, from_=0, to=999, width=5)
        tempo_jogado_edit_minutos = tk.Spinbox(
            janela_edicao, from_=0, to=59, width=5)
        tempo_jogado = jogo_selecionado.get('Tempo Jogado', '').strip()
        if tempo_jogado:
            horas, minutos = map(int, tempo_jogado.split(":"))
        else:
            horas, minutos = 0, 0
        tempo_jogado_edit_horas.delete(0, tk.END)
        tempo_jogado_edit_horas.insert(0, horas)
        tempo_jogado_edit_minutos.delete(0, tk.END)
        tempo_jogado_edit_minutos.insert(0, minutos)

        nota_slider = tk.Scale(janela_edicao, from_=1,
                               to=10, orient=tk.HORIZONTAL)
        nota_slider.set(
            float(jogo_selecionado['Nota']) if jogo_selecionado['Nota'] else 1)

        # Posicionar widgets na janela
        titulo_edit.grid(row=0, column=1, padx=5, pady=5, sticky="EW")
        genero_edit.grid(row=1, column=1, padx=5, pady=5, sticky="EW")
        plataforma_edit.grid(row=2, column=1, padx=5, pady=5, sticky="EW")
        data_zeramento_edit.grid(row=3, column=1, padx=5, pady=5, sticky="EW")
        forma_zeramento_edit.grid(row=4, column=1, padx=5, pady=5, sticky="EW")
        descricao_zeramento_edit.grid(
            row=5, column=1, padx=5, pady=5, sticky="EW")
        tempo_jogado_edit_horas.grid(
            row=6, column=1, padx=5, pady=5, sticky="W")
        tempo_jogado_edit_minutos.grid(
            row=6, column=1, padx=5, pady=5, sticky="E")
        nota_slider.grid(row=7, column=1, padx=5, pady=5, sticky="EW")

        # Função para desativar/ativar campos com base no estado
        def atualizar_campos_edicao(estado):
            if estado in ["Planejo Jogar", "Desistência"]:
                # Desativar campos
                data_zeramento_edit.delete(0, tk.END)
                data_zeramento_edit.config(state="disabled")
                tempo_jogado_edit_horas.delete(0, tk.END)
                tempo_jogado_edit_horas.insert(0, 0)
                tempo_jogado_edit_horas.config(state="disabled")
                tempo_jogado_edit_minutos.delete(0, tk.END)
                tempo_jogado_edit_minutos.insert(0, 0)
                tempo_jogado_edit_minutos.config(state="disabled")
                nota_slider.set(1)
                nota_slider.config(state="disabled")
            else:
                # Reativar campos
                data_zeramento_edit.config(state="normal")
                tempo_jogado_edit_horas.config(state="normal")
                tempo_jogado_edit_minutos.config(state="normal")
                nota_slider.config(state="normal")

        # Chamar a função imediatamente com o estado atual
        atualizar_campos_edicao(forma_zeramento_edit.get())

        # Atualizar campos ao alterar o estado
        forma_zeramento_edit.bind(
            "<<ComboboxSelected>>", lambda event: atualizar_campos_edicao(forma_zeramento_edit.get()))

        # Botão para salvar alterações
        def salvar_edicao():
            try:
                jogo_selecionado['Título'] = titulo_edit.get()
                jogo_selecionado['Gênero'] = genero_edit.get()
                jogo_selecionado['Plataforma'] = plataforma_edit.get()
                jogo_selecionado['Data de Zeramento'] = data_zeramento_edit.get(
                ) if data_zeramento_edit["state"] == "normal" else ""
                jogo_selecionado['Forma de Zeramento'] = forma_zeramento_edit.get()
                jogo_selecionado['Descrição de Zeramento'] = descricao_zeramento_edit.get(
                )
                if tempo_jogado_edit_horas["state"] == "normal":
                    horas = int(tempo_jogado_edit_horas.get())
                    minutos = int(tempo_jogado_edit_minutos.get())
                    jogo_selecionado['Tempo Jogado'] = f"{horas}:{minutos:02d}"
                else:
                    jogo_selecionado['Tempo Jogado'] = ""
                jogo_selecionado['Nota'] = nota_slider.get(
                ) if nota_slider["state"] == "normal" else None

                messagebox.showinfo("Sucesso", "Jogo editado com sucesso!")
                atualizar_lista()
                janela_edicao.destroy()
            except Exception as e:
                messagebox.showerror(
                    "Erro", f"Erro ao salvar as alterações: {e}")

        salvar_button = tk.Button(janela_edicao, text="Salvar Edição",
                                  command=salvar_edicao)
        salvar_button.grid(row=8, column=0, columnspan=2, pady=10)
        estilizar_botao(salvar_button, cor_fundo="gray", largura=20, altura=1)

        excluir_button = tk.Button(janela_edicao, text="Excluir Jogo", 
                                   command=lambda: [excluir_jogo(), janela_edicao.destroy()])
        excluir_button.grid(row=9, column=0, columnspan=2, pady=10)
        estilizar_botao(excluir_button, cor_fundo="gray", largura=20, altura=1)


def mostrar_jogos_filtrados():
    global filtro_window, filtro_titulo_entry, filtro_genero_entry, filtro_plataforma_combobox
    global filtro_min_nota_entry, filtro_max_nota_entry, filtro_ano_entry, filtro_metodo_combobox

    filtro_window = tk.Toplevel(root)
    filtro_window.title("Filtrar Jogos")
    filtro_window.geometry("420x370")
    filtro_window.resizable(False, False)
    centralizar_janela(filtro_window, 420, 370)

    # Rótulos e campos para os filtros
    filtros = [
        ("Título", 0, tk.Entry),
        ("Gênero", 1, ttk.Combobox),
        ("Plataforma", 2, ttk.Combobox),
        ("Nota mínima", 3, tk.Entry),
        ("Ano", 4, tk.Entry),
        ("Método de Zeramento", 5, ttk.Combobox)
    ]

    header_label = tk.Label(filtro_window, text="Selecione os Filtros", font=(
        "Arial", 14, "bold"), fg="#333")
    header_label.grid(row=0, column=0, columnspan=2, pady=(10, 20))

    filtro_titulo_entry = tk.Entry(filtro_window, width=30)
    filtro_genero_entry = ttk.Combobox(
        filtro_window, values=generos_disponiveis, width=27)
    filtro_plataforma_combobox = ttk.Combobox(
        filtro_window, values=plataformas_disponiveis, width=27)
    filtro_min_nota_entry = tk.Entry(filtro_window, width=30)
    filtro_ano_entry = tk.Entry(filtro_window, width=30)
    filtro_metodo_combobox = ttk.Combobox(filtro_window, values=[
                                          "História", "100%", "Platina", "Planejo Jogar", "Desistência"], width=27)

    widgets = [
        ("Título", filtro_titulo_entry),
        ("Gênero", filtro_genero_entry),
        ("Plataforma", filtro_plataforma_combobox),
        ("Nota mínima", filtro_min_nota_entry),
        ("Ano", filtro_ano_entry),
        ("Método de Zeramento", filtro_metodo_combobox)
    ]

    for idx, (label_text, widget) in enumerate(widgets, start=1):
        label = tk.Label(filtro_window, text=label_text,
                         font=("Arial", 10, "bold"))
        label.grid(row=idx, column=0, padx=20, pady=10, sticky="E")
        widget.grid(row=idx, column=1, padx=20, pady=10, sticky="W")

    # Função para aplicar o filtro
    def aplicar_filtro():
        global jogos_filtrados
        filtros = {
            "Título": filtro_titulo_entry.get(),
            "Gênero": filtro_genero_entry.get(),
            "Plataforma": filtro_plataforma_combobox.get(),
            "Nota": filtro_min_nota_entry.get(),
            "Ano": filtro_ano_entry.get(),
            "Método": filtro_metodo_combobox.get()
        }

        def validar_jogo(jogo):
            if filtros["Título"] and filtros["Título"].lower() not in jogo["Título"].lower():
                return False
            if filtros["Gênero"] and filtros["Gênero"] != jogo["Gênero"]:
                return False
            if filtros["Plataforma"] and filtros["Plataforma"] != jogo["Plataforma"]:
                return False
            if filtros["Nota"]:
                try:
                    if float(jogo["Nota"]) < float(filtros["Nota"]):
                        return False
                except ValueError:
                    messagebox.showerror(
                        "Erro", "Nota inválida. Use números entre 1 e 10.")
                    return False
            if filtros["Ano"]:
                try:
                    ano_jogo = datetime.strptime(
                        jogo["Data de Zeramento"], "%d/%m/%Y").year
                    if int(filtros["Ano"]) != ano_jogo:
                        return False
                except ValueError:
                    messagebox.showerror("Erro", "Ano inválido.")
                    return False
            if filtros["Método"] and filtros["Método"] != jogo["Forma de Zeramento"]:
                return False
            return True

        jogos_filtrados = [jogo for jogo in lista_jogos if validar_jogo(jogo)]

        atualizar_lista(jogos_filtrados)
        filtro_window.destroy()

    botao_filtro = tk.Button(
        filtro_window, text="Aplicar Filtro", command=aplicar_filtro,
        bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=15
    )
    botao_filtro.grid(row=7, column=0, columnspan=2, pady=(10, 5))

    estilizar_botao(botao_filtro, cor_fundo="gray", largura=15, altura=1)


def gerenciar_checklist():
    pasta_saves = "saves"
    os.makedirs(pasta_saves, exist_ok=True)
    caminho_arquivo = os.path.join(pasta_saves, "tarefas.json")

    checklist_window = tk.Toplevel(root)
    checklist_window.title("Tarefas")
    checklist_window.geometry("500x500")
    checklist_window.resizable(False, False)
    centralizar_janela(checklist_window, 500, 500)

    try:
        with open(caminho_arquivo, "r") as arquivo:
            tarefas = json.load(arquivo)
    except (FileNotFoundError, json.JSONDecodeError):
        tarefas = []

    # Salvar tarefas no arquivo JSON
    def salvar_tarefas():
        with open(caminho_arquivo, "w") as arquivo:
            json.dump(tarefas, arquivo, indent=4)

    def atualizar_tarefas():
        tarefas_listbox.delete(0, tk.END)
        for tarefa in tarefas:
            status = "✔" if all(m["concluido"]
                                for m in tarefa["missoes"]) else "✘"
            tarefas_listbox.insert(tk.END, f"{status} {tarefa['nome']}")

    def adicionar_tarefa():
        nova_tarefa_nome = tarefa_entry.get()
        if nova_tarefa_nome:
            tarefas.append({"nome": nova_tarefa_nome, "missoes": []})
            salvar_tarefas()
            atualizar_tarefas()
            tarefa_entry.delete(0, tk.END)

    def excluir_tarefa():
        selecionado = tarefas_listbox.curselection()
        if not selecionado:
            messagebox.showerror("Erro", "Selecione uma tarefa para excluir.")
            return

        indice = selecionado[0]
        confirmar = messagebox.askyesno(
            "Confirmação", "Tem certeza que deseja excluir esta tarefa?")
        if confirmar:
            tarefas.pop(indice)
            salvar_tarefas()
            atualizar_tarefas()

    def gerenciar_missoes():
        selecionado = tarefas_listbox.curselection()
        if not selecionado:
            messagebox.showerror(
                "Erro", "Selecione uma tarefa para gerenciar.")
            return

        indice_tarefa = selecionado[0]
        tarefa = tarefas[indice_tarefa]

        missoes_window = tk.Toplevel(checklist_window)
        missoes_window.title(f"Missões - {tarefa['nome']}")
        missoes_window.geometry("400x450")
        missoes_window.resizable(False, False)
        centralizar_janela(missoes_window, 400, 450)

        progresso_label = tk.Label(
            missoes_window, text="", font=("Arial", 10, "bold"))
        progresso_label.pack(pady=5)

        missoes_listbox = tk.Listbox(missoes_window, width=40, height=15)
        missoes_listbox.pack(padx=10, pady=10)

        def atualizar_missoes():
            missoes_listbox.delete(0, tk.END)
            concluido = sum(1 for m in tarefa["missoes"] if m["concluido"])
            total = len(tarefa["missoes"])
            progresso_label.config(
                text=f"Progresso: {concluido}/{total} missões concluídas"
            )

            for missao in tarefa["missoes"]:
                status = "✔" if missao["concluido"] else "✘"
                missoes_listbox.insert(tk.END, f"{status} {missao['nome']}")

            salvar_tarefas()
            atualizar_tarefas()

        def adicionar_missao():
            nova_missao_nome = missao_entry.get()
            if nova_missao_nome:
                tarefa["missoes"].append(
                    {"nome": nova_missao_nome, "concluido": False})
                salvar_tarefas()
                atualizar_missoes()
                missao_entry.delete(0, tk.END)

        def excluir_missao():
            selecionado = missoes_listbox.curselection()
            if not selecionado:
                messagebox.showerror(
                    "Erro", "Selecione uma missão para excluir.")
                return

            indice_missao = selecionado[0]
            confirmar = messagebox.askyesno(
                "Confirmação", "Tem certeza que deseja excluir esta missão?")
            if confirmar:
                tarefa["missoes"].pop(indice_missao)
                salvar_tarefas()
                atualizar_missoes()

        def alternar_status_missao():
            selecionado = missoes_listbox.curselection()
            if not selecionado:
                return
            indice_missao = selecionado[0]
            tarefa["missoes"][indice_missao]["concluido"] = not tarefa["missoes"][indice_missao]["concluido"]
            salvar_tarefas()
            atualizar_missoes()

        missao_entry = tk.Entry(missoes_window, width=30)
        missao_entry.pack(pady=3)
        adicionar_missao_button = tk.Button(
            missoes_window, text="Adicionar Missão", command=adicionar_missao, bg="#4CAF50", fg="white"
        )
        estilizar_botao(adicionar_missao_button,
                        cor_fundo="#4CAF50", largura=15, altura=1)
        adicionar_missao_button.pack(pady=5)

        excluir_missao_button = tk.Button(
            missoes_window, text="Excluir Missão", command=excluir_missao, bg="#f44336", fg="white"
        )
        estilizar_botao(excluir_missao_button,
                        cor_fundo="#f44336", largura=15, altura=1)
        excluir_missao_button.pack(pady=5)

        alternar_status_button = tk.Button(
            missoes_window, text="Marcar/Desmarcar", command=alternar_status_missao, bg="#FFD700", fg="black"
        )
        estilizar_botao(alternar_status_button,
                        cor_fundo="gray", largura=15, altura=1)
        alternar_status_button.pack(pady=5)

        atualizar_missoes()

    # Layout Principal
    header_label = tk.Label(
        checklist_window, text="Checklist de Tarefas", font=("Arial", 16, "bold"))
    header_label.pack(pady=10)

    frame_lista = tk.Frame(checklist_window)
    frame_lista.pack(pady=10, fill="both", expand=True)

    tarefas_listbox = tk.Listbox(frame_lista, width=50, height=15)
    tarefas_listbox.pack(side="left", fill="both", expand=True, padx=10)

    scrollbar = ttk.Scrollbar(
        frame_lista, orient="vertical", command=tarefas_listbox.yview)
    tarefas_listbox.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    frame_acoes = tk.Frame(checklist_window)
    frame_acoes.pack(pady=10)

    tk.Label(frame_acoes, text="Nova Tarefa:").grid(
        row=0, column=0, padx=5, pady=5, sticky="w")
    tarefa_entry = tk.Entry(frame_acoes, width=30)
    tarefa_entry.grid(row=0, column=1, padx=5, pady=5)

    # Aplicando nos botões
    adicionar_button = tk.Button(
        frame_acoes, text="Adicionar", command=adicionar_tarefa)
    adicionar_button["bg"] = "gray"
    adicionar_button["fg"] = "white"
    estilizar_botao(adicionar_button, "gray", largura=15, altura=1)
    adicionar_button.grid(row=0, column=2, padx=5, pady=5)

    excluir_button = tk.Button(
        frame_acoes, text="Excluir", command=excluir_tarefa)
    estilizar_botao(excluir_button, "#f44336", largura=15, altura=1)
    excluir_button.grid(row=1, column=2, padx=5, pady=5)

    gerenciar_missoes_button = tk.Button(
        frame_acoes, text="Gerenciar Missões", command=gerenciar_missoes)
    estilizar_botao(gerenciar_missoes_button, "#808080", largura=15, altura=1)
    gerenciar_missoes_button.grid(row=2, column=0, columnspan=3, pady=10)

    atualizar_tarefas()


def mostrar_informacoes(event):
    selecionado = lista_jogos_listbox.curselection()
    if selecionado:
        indice_selecionado = int(selecionado[0])
        if 0 <= indice_selecionado < len(jogos_filtrados):
            # Obter o jogo diretamente da lista filtrada
            jogo_selecionado = jogos_filtrados[indice_selecionado]

            mensagem = f"Informações do Jogo:\n\nTítulo: {jogo_selecionado['Título']}\nGênero: {jogo_selecionado['Gênero']}\nPlataforma: {jogo_selecionado['Plataforma']}\nData de Zeramento: {jogo_selecionado['Data de Zeramento']}\nForma de Zeramento: {jogo_selecionado['Forma de Zeramento']}\nDescrição de Zeramento: {jogo_selecionado['Descrição de Zeramento']}\nTempo Jogado: {jogo_selecionado['Tempo Jogado']}\nNota: {jogo_selecionado['Nota']}"
            messagebox.showinfo("Informações do Jogo", mensagem)


def copiar_nome():
    # Obtenha o índice do jogo selecionado
    indice_selecionado = lista_jogos_listbox.curselection()[0]
    jogo_selecionado = lista_jogos_listbox.get(indice_selecionado)

    # Use uma expressão regular para encontrar o título do jogo após o número, ponto e emoji
    padrao_titulo = r'\d+\.\s*[^\w\s]+\s*(.+)'
    correspondencia = re.search(padrao_titulo, jogo_selecionado)
    if correspondencia:
        titulo_do_jogo = correspondencia.group(1)
    else:
        titulo_do_jogo = "Título não encontrado"

    # Copie o título do jogo para a área de transferência
    pyperclip.copy(titulo_do_jogo)

def selecionar_wallpaper():
    """Abre um seletor de arquivos para escolher a imagem e inicia a edição."""
    caminho_imagem = filedialog.askopenfilename(filetypes=[("Imagens", "*.png;*.jpg;*.jpeg")])
    if caminho_imagem:
        editar_wallpaper(caminho_imagem)


def editar_wallpaper(caminho_imagem):
    """Abre uma janela para recortar o wallpaper antes de salvar."""
    global img_tk, img_editando, canvas, rect_id, img_original
    global x_inicial, y_inicial, desloc_x, desloc_y, redimensionando, largura_recorte, altura_recorte

    # Proporção fixa do recorte
    largura_recorte = 600
    altura_recorte = 400

    # Criar a janela de edição
    janela_edicao = tk.Toplevel(root)
    janela_edicao.title("Editar Wallpaper")
    janela_edicao.geometry("850x650")
    janela_edicao.configure(bg="#2C3E50")
    centralizar_janela(janela_edicao, 850, 650)
    janela_edicao.resizable(False, False)

    # Carregar a imagem e redimensioná-la
    img_original = Image.open(caminho_imagem)
    fator_escala = min(800 / img_original.width, 550 / img_original.height)
    nova_largura = int(img_original.width * fator_escala)
    nova_altura = int(img_original.height * fator_escala)
    img_editando = img_original.resize((nova_largura, nova_altura), Image.LANCZOS)

    # Converter para exibição no Tkinter
    img_tk = ImageTk.PhotoImage(img_editando)

    # Criar o Canvas
    canvas = tk.Canvas(janela_edicao, width=nova_largura, height=nova_altura, bg="#ECF0F1", bd=2, relief="solid")
    canvas.pack(pady=10)

    # Exibir a imagem no Canvas
    canvas.create_image(0, 0, anchor="nw", image=img_tk)

    # Posicionamento inicial do recorte (centralizado)
    x_inicial = (nova_largura - largura_recorte) // 2
    y_inicial = (nova_altura - altura_recorte) // 2
    x_final = x_inicial + largura_recorte
    y_final = y_inicial + altura_recorte

    # Criar o retângulo de recorte já visível
    rect_id = canvas.create_rectangle(x_inicial, y_inicial, x_final, y_final, outline="red", width=2, tags="recorte")

    # Variáveis de controle
    desloc_x, desloc_y = 0, 0
    redimensionando = False

    # Função para iniciar o movimento do recorte
    def iniciar_movimento(event):
        global desloc_x, desloc_y
        desloc_x, desloc_y = event.x, event.y

    # Função para mover o recorte
    def mover_recorte(event):
        global desloc_x, desloc_y
        dx = event.x - desloc_x
        dy = event.y - desloc_y
        desloc_x, desloc_y = event.x, event.y
        canvas.move(rect_id, dx, dy)

    # Função para iniciar o redimensionamento
    def iniciar_redimensionamento(event):
        global desloc_x, desloc_y, redimensionando
        desloc_x, desloc_y = event.x, event.y
        redimensionando = True

    # Função para redimensionar mantendo a proporção
    def redimensionar_recorte(event):
        global largura_recorte, altura_recorte, x_inicial, y_inicial
        if redimensionando:
            nova_largura = max(100, event.x - x_inicial)  # Mínimo de 100 px
            nova_altura = int(nova_largura * (altura_recorte / largura_recorte))  # Mantém a proporção

            x_final = x_inicial + nova_largura
            y_final = y_inicial + nova_altura

            # Atualiza o retângulo mantendo a proporção 600x400
            canvas.coords(rect_id, x_inicial, y_inicial, x_final, y_final)

    # Vincular eventos para mover e redimensionar
    canvas.tag_bind("recorte", "<ButtonPress-1>", iniciar_movimento)
    canvas.tag_bind("recorte", "<B1-Motion>", mover_recorte)
    canvas.bind("<ButtonPress-3>", iniciar_redimensionamento)  # Clique direito para redimensionar
    canvas.bind("<B3-Motion>", redimensionar_recorte)  # Arrastar para redimensionar

    # Função para salvar a imagem recortada
    def salvar_recorte():
        x1, y1, x2, y2 = canvas.coords(rect_id)
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

        if x2 > x1 and y2 > y1:
            img_crop = img_editando.crop((x1, y1, x2, y2))
            img_crop = img_crop.resize((largura_recorte, altura_recorte))
            salvar_wallpaper(img_crop, largura_recorte, altura_recorte)
        else:
            messagebox.showerror("Erro", "Selecione uma área válida para recorte.")
        
        atualizar_wallpaper()
        carregar_background()
        janela_edicao.destroy()
    # Botão para salvar
    btn_confirmar = tk.Button(janela_edicao, text="Salvar Recorte", command=salvar_recorte, bg="#27AE60", fg="white",
                              font=("Arial", 12, "bold"), padx=10, pady=5, relief="raised", borderwidth=3)
    btn_confirmar.pack(pady=10)

def salvar_wallpaper(imagem, largura, altura):
    """Salva a imagem cortada como wallpaper.png na pasta layout."""
    pasta_layout = "layout"
    os.makedirs(pasta_layout, exist_ok=True)

    caminho_salvar = os.path.join(pasta_layout, "wallpaper.png")
    imagem_redimensionada = imagem.resize((largura, altura))
    imagem_redimensionada.save(caminho_salvar, "PNG")

    messagebox.showinfo("Sucesso!", "Wallpaper atualizado com sucesso!")

    # Atualizar a interface
    atualizar_wallpaper()

def atualizar_wallpaper():
    caminho_wallpaper = os.path.join("layout", "wallpaper.png")
    
    if os.path.exists(caminho_wallpaper):
        img = Image.open(caminho_wallpaper)
        img_tk = ImageTk.PhotoImage(img)

        # Aplicar o wallpaper como fundo
        wallpaper_label.config(image=img_tk)
        wallpaper_label.image = img_tk

def carregar_wallpaper():
    global wallpaper_tk

    caminho_wallpaper = os.path.join("layout", "wallpaper.png")
    
    if os.path.exists(caminho_wallpaper):
        img = Image.open(caminho_wallpaper).convert("RGBA")
        wallpaper_tk = ImageTk.PhotoImage(img)

        # Aplicar o wallpaper como fundo
        wallpaper_label.config(image=wallpaper_tk)
        wallpaper_label.image = wallpaper_tk  # Evita descarte

def carregar_background():
    global background_tk

    caminho_wallpaper = os.path.join("layout", "wallpaper.png")
    caminho_background = os.path.join("layout", "background.png")

    if os.path.exists(caminho_wallpaper) and os.path.exists(caminho_background):
        wallpaper = Image.open(caminho_wallpaper).convert("RGBA")
        background = Image.open(caminho_background).convert("RGBA")

        # Garantir que ambas as imagens tenham o mesmo tamanho
        wallpaper = wallpaper.resize(background.size)

        # Mesclar as imagens (background sobre wallpaper)
        imagem_final = Image.alpha_composite(wallpaper, background)

        # Converter para o formato do Tkinter
        background_tk = ImageTk.PhotoImage(imagem_final)

        # Aplicar ao Label
        background_label.config(image=background_tk)
        background_label.image = background_tk  # Evita descarte

# -------------------------------
# OUTRAS UTILIDADES
# -------------------------------


def show_error_message(message):
    error_window = tk.Toplevel(root)
    error_window.title("Algo está errado ;-;")
    error_label = tk.Label(error_window, text=message)
    error_label.pack()
    ok_button = tk.Button(error_window, text="OK",
                          command=error_window.destroy)
    ok_button.pack()


def on_closing():
    resposta = messagebox.askyesnocancel(
        "Sair", "Deseja salvar antes de Sair? Os dados não salvos serão perdidos.")
    if resposta is None:
        return
    elif resposta:
        salvar_lista()
    root.destroy()

def criar_aba_resumo():
    resumo_window = tk.Toplevel(root)
    resumo_window.title("Resumo Geral")
    resumo_window.geometry("590x650")
    resumo_window.resizable(False, False)
    centralizar_janela(resumo_window, 590, 650)

    # Filtrar os jogos por estado
    jogos_zerados = [
        jogo for jogo in lista_jogos if jogo.get("Data de Zeramento")]
    jogos_desistidos = [jogo for jogo in lista_jogos if jogo.get(
        "Forma de Zeramento") == "Desistência"]
    jogos_planejados = [jogo for jogo in lista_jogos if jogo.get(
        "Forma de Zeramento") == "Planejo Jogar"]

    # Frame principal com Canvas e Scrollbar
    canvas = tk.Canvas(resumo_window, bg="black", width=780)
    scrollbar = ttk.Scrollbar(
        resumo_window, orient="vertical", command=canvas.yview)
    scroll_frame = tk.Frame(canvas, bg="black")

    scroll_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Estilo de texto
    header_font = tkFont.Font(family="Arial", size=16, weight="bold")
    text_font = tkFont.Font(family="Arial", size=14)

    # Cabeçalho
    tk.Label(scroll_frame, text="Resumo Geral", font=header_font,
             bg="black", fg="white").pack(pady=10)

    # Total de jogos zerados, desistidos e planejados
    tk.Label(scroll_frame, text=f"Total de Jogos Zerados: {len(jogos_zerados)}",
             font=text_font, bg="black", fg="white").pack(pady=5)
    tk.Label(scroll_frame, text=f"Total de Jogos Desistidos: {len(jogos_desistidos)}",
             font=text_font, bg="black", fg="white").pack(pady=5)
    tk.Label(scroll_frame, text=f"Total de Jogos Planejados: {len(jogos_planejados)}",
             font=text_font, bg="black", fg="white").pack(pady=5)

    # Média de notas
    notas = [float(jogo["Nota"]) for jogo in jogos_zerados if jogo.get("Nota")]
    media_notas = sum(notas) / len(notas) if notas else 0
    tk.Label(
        scroll_frame,
        text=f"Média de Notas dos Jogos Zerados: {media_notas:.2f}",
        font=text_font,
        bg="black",
        fg="white",
    ).pack(pady=5)

    # Resumo por Gêneros
    tk.Label(
        scroll_frame,
        text="Resumo por Gêneros (Jogos e Tempo):",
        font=header_font,
        bg="black",
        fg="white",
    ).pack(pady=10)

    genero_contagem = {}
    for jogo in jogos_zerados:
        genero = jogo.get("Gênero", "Outro")
        genero_contagem[genero] = genero_contagem.get(
            genero, {"quantidade": 0, "tempo": 0})
        genero_contagem[genero]["quantidade"] += 1
        genero_contagem[genero]["tempo"] += calcular_total_minutos(
            jogo.get("Tempo Jogado", "0:00"))

    generos_frame = tk.Frame(scroll_frame, bg="black")
    generos_frame.pack(pady=5, fill="x")

    generos_tree = ttk.Treeview(generos_frame, columns=(
        "Genero", "Quantidade", "Tempo"), show="headings", height=5)
    generos_tree.heading("Genero", text="Gênero")
    generos_tree.heading("Quantidade", text="Quantidade")
    generos_tree.heading("Tempo", text="Tempo Jogado (Horas)")
    generos_tree.column("Genero", width=200, anchor="center")
    generos_tree.column("Quantidade", width=100, anchor="center")
    generos_tree.column("Tempo", width=150, anchor="center")

    for genero, dados in genero_contagem.items():
        horas = dados["tempo"] // 60
        generos_tree.insert("", "end", values=(
            genero, dados["quantidade"], f"{horas}h"))
    generos_tree.pack(side="left", fill="both", expand=True)

    generos_scrollbar = ttk.Scrollbar(
        generos_frame, orient="vertical", command=generos_tree.yview)
    generos_tree.configure(yscrollcommand=generos_scrollbar.set)
    generos_scrollbar.pack(side="right", fill="y")

    # Resumo por Estado e Plataforma
    tk.Label(
        scroll_frame,
        text="Resumo por Estado (Plataformas):",
        font=header_font,
        bg="black",
        fg="white",
    ).pack(pady=10)

    plataforma_resumo = {}
    for jogo in lista_jogos:
        plataforma = jogo["Plataforma"]
        if plataforma not in plataforma_resumo:
            plataforma_resumo[plataforma] = {
                "Planeja Jogar": 0, "Desistidos": 0, "Zerados": 0, "Horas Jogadas": 0}

        estado = jogo["Forma de Zeramento"]
        if estado == "Planejo Jogar":
            plataforma_resumo[plataforma]["Planeja Jogar"] += 1
        elif estado == "Desistência":
            plataforma_resumo[plataforma]["Desistidos"] += 1
        elif estado in ["História", "100%", "Platina"]:
            plataforma_resumo[plataforma]["Zerados"] += 1
            if jogo.get("Tempo Jogado"):
                plataforma_resumo[plataforma]["Horas Jogadas"] += calcular_total_minutos(
                    jogo["Tempo Jogado"]) // 60

    plataformas_frame = tk.Frame(scroll_frame, bg="black")
    plataformas_frame.pack(pady=5, fill="x")

    plataformas_tree = ttk.Treeview(
        plataformas_frame,
        columns=("Plataforma", "Planeja Jogar",
                 "Desistidos", "Zerados", "Horas Jogadas"),
        show="headings",
        height=5
    )
    plataformas_tree.heading("Plataforma", text="Plataforma")
    plataformas_tree.heading("Planeja Jogar", text="Planeja Jogar")
    plataformas_tree.heading("Desistidos", text="Desistidos")
    plataformas_tree.heading("Zerados", text="Zerados")
    plataformas_tree.heading("Horas Jogadas", text="Horas Jogadas (h)")
    plataformas_tree.column("Plataforma", width=150, anchor="center")
    plataformas_tree.column("Planeja Jogar", width=100, anchor="center")
    plataformas_tree.column("Desistidos", width=100, anchor="center")
    plataformas_tree.column("Zerados", width=100, anchor="center")
    plataformas_tree.column("Horas Jogadas", width=120, anchor="center")

    for plataforma, stats in plataforma_resumo.items():
        plataformas_tree.insert("", "end", values=(
            plataforma,
            stats["Planeja Jogar"],
            stats["Desistidos"],
            stats["Zerados"],
            f"{stats['Horas Jogadas']}h"
        ))
    plataformas_tree.pack(side="left", fill="both", expand=True)

    plataformas_scrollbar = ttk.Scrollbar(
        plataformas_frame, orient="vertical", command=plataformas_tree.yview)
    plataformas_tree.configure(yscrollcommand=plataformas_scrollbar.set)
    plataformas_scrollbar.pack(side="right", fill="y")

    # Tempo total jogado
    total_minutos = sum(calcular_total_minutos(
        jogo["Tempo Jogado"]) for jogo in jogos_zerados if jogo.get("Tempo Jogado"))
    total_horas = total_minutos // 60
    total_dias = total_horas // 24
    tk.Label(
        scroll_frame,
        text=f"Tempo Total Jogado: {total_dias} dias, {total_horas % 24} horas",
        font=text_font,
        bg="black",
        fg="white",
    ).pack(pady=10)


def substituir_espaco_por_dois_pontos(event):
    tempo_jogado_entry_var.set(tempo_jogado_entry_var.get().replace(" ", ":"))


# -------------------------------
# PROGRAMA PRINCIPAL
# -------------------------------

# Inicializa a lista de jogos
lista_jogos = carregar_lista()
titulo_entry = None
genero_entry = None
plataforma_entry = None
filtro_window = None
nota_entry = None
janela_edicao = None
jogos_filtrados = lista_jogos.copy()
datetime_value = "algum_valor"


root = tk.Tk()
root.title("Registro ULTIMATE de jogos")


icon_path = os.path.join(os.getcwd(), "layout", "icon.ico")
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)
else:
    print(f"Ícone não encontrado no caminho: {icon_path}")


root.protocol("WM_DELETE_WINDOW", on_closing)

# Caminho do arquivo na pasta 'layout'
image_path = os.path.join(os.getcwd(), "layout", "Background.png")

try:
    # Criar label para o wallpaper
    wallpaper_label = tk.Label(root)
    wallpaper_label.place(x=0, y=0, relwidth=1, relheight=1)

    # Carregar o background garantindo transparência
    img_background = Image.open(image_path).convert("RGBA")
    background_tk = ImageTk.PhotoImage(img_background)

    # Criar label para o background com transparência
    background_label = tk.Label(root, image=background_tk)
    background_label.place(x=0, y=0, relwidth=1, relheight=1)

    # Manter referência da imagem para evitar descarte
    background_label.image = background_tk

except Exception as e:
    messagebox.showerror("Erro", f"Não foi possível carregar o fundo.\nDetalhes: {e}")

# Atualizar as imagens de fundo
atualizar_wallpaper()
carregar_background()

titulo_label = tk.Label(root, text="Título*:")
titulo_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
titulo_entry = tk.Entry(root)
titulo_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)

genero_entry = tk.Entry(root)
genero_entry.grid(row=1, column=2, padx=10, pady=5, sticky=tk.W)

genero_label = tk.Label(root, text="Gênero*:")
genero_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
generos_disponiveis = [
    # RPG
    "RPG", "Action RPG", "RPG de Turno", "RPG Tático", "JRPG", "Action JRPG",
    "JRPG de Turno", "JRPG Tático", "Dungeon Crawler", "MMORPG", "True RPG",

    # Aventura
    "Aventura", "Aventura Gráfica", "Point and Click", "Metroidvania", "Survival Horror",

    # Ação
    "Ação", "Hack and Slash", "Beat 'em Up", "Stealth", "Action-Adventure",

    # Estratégia
    "Estratégia", "RTS (Real-Time Strategy)", "TBS (Turn-Based Strategy)",
    "Tower Defense", "4X (eXplore, eXpand, eXploit, eXterminate)",

    # Simulação
    "Simulação", "Simulação de Vida", "Simulação de Construção", "Simulação de Negócios",
    "Simulação de Voo", "Simulação de Veículos", "Simulação Social",

    # Esportes
    "Esportes", "Futebol", "Basquete", "Corrida", "Golfe", "Tênis", "Esportes Radicais",
    "Automobilismo", "Futebol Americano",

    # Puzzle
    "Puzzle", "Quebra-Cabeças Lógicos", "Match-3", "Jogos de Palavras", "Sokoban", "Escape Room",

    # Luta
    "Luta", "2D Fighting", "3D Fighting", "Arena Fighting", "Party Fighting", "Beat 'em Up",

    # Tiro
    "Tiro", "FPS (First-Person Shooter)", "TPS (Third-Person Shooter)", "Shoot 'em Up",
    "Light Gun Shooter", "Bullet Hell",

    # Horror
    "Horror", "Survival Horror", "Psychological Horror", "Action Horror", "VR Horror",

    # Sandbox
    "Sandbox", "World Builder", "Exploration", "Open World", "Criativo",

    # Party Games
    "Jogos de Festa", "Minigames", "Quiz", "Jogos de Tabuleiro Adaptados",

    # Outros Gêneros
    "Educação", "Treinamento", "Documentário", "Outro"
]
genero_var = tk.StringVar()
genero_combobox = ttk.Combobox(
    root, textvariable=genero_var, values=generos_disponiveis, state="readonly")
genero_combobox.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
genero_combobox.set("")
genero_combobox.bind("<KeyRelease>", atualizar_entry)
genero_combobox.bind("<<ComboboxSelected>>", atualizar_genero)

plataformas_disponiveis = [
    # Atari
    "Atari 2600", "Atari 5200", "Atari 7800",

    # Nintendo
    "NES (Nintendo Entertainment System)", "SNES (Super Nintendo Entertainment System)",
    "Nintendo 64", "GameCube", "Game Boy", "Game Boy Color",
    "Game Boy Advance", "Nintendo DS", "Nintendo Switch",

    # Sega
    "Sega Master System", "Sega Genesis (Mega Drive)",
    "Sega Saturn", "Sega Dreamcast", "Game Gear",

    # Sony
    "PlayStation 1", "PlayStation 2", "PlayStation 3",
    "PlayStation 4", "PlayStation 5", "PlayStation Portable",

    # Microsoft
    "Xbox Clássico", "Xbox 360", "Xbox One", "Xbox Series X|S",

    # SNK
    "Neo Geo", "Neo Geo Pocket", "Neo Geo Pocket Color",

    # NEC
    "TurboGrafx-16 (PC Engine)", "TurboGrafx-CD",

    # Mattel
    "Intellivision",

    # Coleco
    "ColecoVision",

    # Commodore
    "Commodore 64", "Amiga",

    # Sinclair
    "ZX Spectrum",

    # Panasonic/GoldStar
    "3DO",

    # Modernos e genéricos
    "PC", "Mobile",

    # Outros
    "Outro"
]

# Campo de entrada e label para plataforma
plataforma_label = tk.Label(root, text="Plataforma*:")
plataforma_label.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)

# Variável para armazenar o valor da plataforma
plataforma_var = tk.StringVar()

# Combobox para plataformas
plataforma_combobox = ttk.Combobox(
    root, textvariable=plataforma_var, values=plataformas_disponiveis, state="readonly"
)
plataforma_combobox.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

# Entrada de texto para personalização
plataforma_entry = tk.Entry(root)
plataforma_entry.grid(row=2, column=2, padx=10, pady=5, sticky=tk.W)

# Vincular a ação de seleção do Combobox à função de atualização
plataforma_combobox.bind("<<ComboboxSelected>>", atualizar_plataforma)

data_zeramento_label = tk.Label(root, text="Data de Zeramento:*")
data_zeramento_label.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
data_zeramento_entry = tk.Entry(root)
data_zeramento_entry.grid(row=3, column=1, padx=10, pady=5, sticky=tk.W)

data_zeramento_entry.bind("<KeyRelease>", formatar_data_zeramento)
data_zeramento_entry.bind("<Double-Button-1>", atualizar_data_zeramento)

forma_zeramento_label = tk.Label(
    root, text="Forma de Zeramento:*", anchor="w")
forma_zeramento_label.grid(row=4, column=0, padx=10, pady=5, sticky=tk.W)
forma_zeramento_var = tk.StringVar()
forma_zeramento_combobox = ttk.Combobox(
    root, textvariable=forma_zeramento_var, values=["História", "100%", "Platina", "Planejo Jogar", "Desistência", "Outro"], state="readonly")
forma_zeramento_combobox.grid(row=4, column=1, padx=10, pady=5, sticky=tk.W)
forma_zeramento_combobox.bind("<<ComboboxSelected>>", atualizar_campos)
forma_zeramento_combobox.set("")

descricao_zeramento_label = tk.Label(
    root, text="Descrição de Zeramento:")
descricao_zeramento_label.grid(row=5, column=0, padx=10, pady=5, sticky=tk.W)
descricao_zeramento_entry = tk.Entry(root)
descricao_zeramento_entry.grid(row=5, column=1, padx=10, pady=5, sticky=tk.W)

# Modificar apenas o campo de entrada de tempo
tempo_label = tk.Label(root, text="Tempo Jogado (H:M)*:")
tempo_label.grid(row=6, column=0, padx=10, pady=5, sticky=tk.W)

tempo_jogado_entry_var = tk.StringVar()
tempo_jogado_entry = tk.Entry(root, textvariable=tempo_jogado_entry_var)
tempo_jogado_entry.grid(row=6, column=1, padx=10, pady=5, sticky=tk.W)

# Vincular a função ao evento de liberação de tecla
tempo_jogado_entry.bind("<KeyRelease>", formatar_tempo_jogado)

nota_label = tk.Label(root, text="Nota:")
nota_label.grid(row=7, column=0, padx=10, pady=5, sticky=tk.W)
nota_slider = tk.Scale(root, from_=1, to=10, orient=tk.HORIZONTAL)
nota_slider.grid(row=7, column=1, padx=10, pady=5, sticky=tk.W)

adicionar_button = tk.Button(
    root, text="Adicionar Jogo", command=adicionar_jogo)
estilizar_botao(adicionar_button, cor_fundo="gray", cor_texto="white",
                largura=15, altura=1, fonte=("Arial", 8, "bold"))
adicionar_button.grid(row=8, column=0, columnspan=2, pady=10)

list_frame = tk.Frame(root)
list_frame.grid(row=0, column=2, rowspan=9, padx=10, pady=5, sticky=tk.N)

lista_jogos_listbox = tk.Listbox(list_frame, width=40, height=15)
lista_jogos_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
scrollbar.config(command=lista_jogos_listbox.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)


lista_jogos_listbox.config(yscrollcommand=scrollbar.set)


lista_jogos_listbox.bind("<Double-Button-1>", mostrar_informacoes)
lista_jogos_listbox.bind("<Button-3>", abrir_menu_contexto)

atualizar_lista()

menu = tk.Menu(root)
root.config(menu=menu)


arquivo_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Arquivo", menu=arquivo_menu)
arquivo_menu.add_command(label="Exportar para PDF", command=exportar_para_pdf)
arquivo_menu.add_command(label="Exportar para Excel",
                         command=exportar_para_excel)
arquivo_menu.add_command(label="Importar de Excel", command=importar_de_excel)
arquivo_menu.add_command(label="Salvar em Arquivo JSON",
                         command=salvar_em_arquivo)
arquivo_menu.add_command(
    label="Carregar de Arquivo JSON", command=carregar_de_arquivo)
arquivo_menu.add_separator()
arquivo_menu.add_command(label="Alterar Wallpaper", command=selecionar_wallpaper)
arquivo_menu.add_separator()
arquivo_menu.add_command(label="Sair", command=on_closing)


editar_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Editar", menu=editar_menu)
editar_menu.add_command(label="Editar Jogo", command=editar_jogo)


filtro_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Filtro", menu=filtro_menu)
filtro_menu.add_command(label="Adicionar Filtro",
                        command=mostrar_jogos_filtrados)
filtro_menu.add_command(label="Limpar Filtros", command=limpar_filtros)

informacoes_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="Informações", menu=informacoes_menu)

# Seção: Resumos e análises gerais
informacoes_menu.add_command(
    label="Tempo Total Jogado", command=calcular_tempo_total_jogado)
informacoes_menu.add_command(
    label="Jogos Mais e Menos longos", command=mostrar_jogos_longos_curto)
informacoes_menu.add_command(
    label="Jogos Zerados (Por Ano)", command=exibir_numero_jogos_zerados_por_ano)
informacoes_menu.add_separator()

# Seção: Análise de notas e plataformas
informacoes_menu.add_command(
    label="Análise de Notas", command=criar_analise_de_notas)
informacoes_menu.add_command(
    label="Jogos por Plataforma", command=criar_distribuicao_plataformas)
informacoes_menu.add_command(
    label="Notas Médias (Por Plataforma)", command=criar_media_notas_plataformas)
informacoes_menu.add_command(
    label="Tempo Jogado (Por Plataforma)", command=criar_tempo_total_plataformas)
informacoes_menu.add_separator()

# Seção: Gráficos e comparações
informacoes_menu.add_command(
    label="Gêneros Mais Jogados (Gráfico)", command=criar_grafico_generos)
informacoes_menu.add_command(
    label="Jogos Zerados ao Longo dos Anos", command=criar_grafico_jogos_por_ano)
informacoes_menu.add_command(
    label="Comparação de Gêneros (Anual)", command=criar_grafico_comparativo_generos)

# Outros comandos fora do menu de informações
menu.add_command(label="Minhas Tarefas", command=gerenciar_checklist)
menu.add_command(label="Resumo da sua Jornada", command=criar_aba_resumo)

largura_janela = 600
altura_janela = 400
centralizar_janela(root, largura_janela, altura_janela)

root.configure(bg="#808080")
root.resizable(False, False)
root.mainloop()
